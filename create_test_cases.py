import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active

# Set title
ws.merge_cells('A1:K1')
ws['A1'] = 'Designing Test Case Sample for Home Service Management System'
ws['A1'].font = Font(bold=True, color='FFFFFF')
ws['A1'].fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
ws['A1'].alignment = Alignment(horizontal='center')

# Set ID row
ws.merge_cells('A2:K2')
ws['A2'] = 'ID: Project_2023'
ws['A2'].font = Font(bold=True, color='FFFFFF')
ws['A2'].fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
ws['A2'].alignment = Alignment(horizontal='center')

# Headers
headers = ['Test Case ID', 'Created By', 'Test Case Description', 'Reviewed By', 'Version']
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

# Test cases data
test_cases = [
    ['TC_001', 'QA Team', 'Test user registration with valid data', 'Senior QA', '1.0'],
    ['TC_002', 'QA Team', 'Test email verification process', 'Senior QA', '1.0'],
    ['TC_003', 'QA Team', 'Test login functionality', 'Senior QA', '1.0'],
    ['TC_004', 'QA Team', 'Test service provider selection', 'Senior QA', '1.0'],
    ['TC_005', 'QA Team', 'Test booking process', 'Senior QA', '1.0'],
    ['TC_006', 'QA Team', 'Test payment processing', 'Senior QA', '1.0'],
    ['TC_007', 'QA Team', 'Test review submission', 'Senior QA', '1.0'],
    ['TC_008', 'QA Team', 'Test profile management', 'Senior QA', '1.0'],
    ['TC_009', 'QA Team', 'Test security features', 'Senior QA', '1.0']
]

# Add test cases
for row, test_case in enumerate(test_cases, start=5):
    for col, value in enumerate(test_case, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

# Adjust column widths
for col in range(1, 6):
    ws.column_dimensions[get_column_letter(col)].width = 20

# Save the workbook
wb.save('Test_Cases.xlsx') 